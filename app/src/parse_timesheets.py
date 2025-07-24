from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import List
from dataclasses import dataclass
from datetime import datetime, date, timedelta
from pclaw import DH_fill_time_entry
from config import *

@dataclass
class TimeEntry:
    date: date
    client: str
    matter: str
    description: str
    time_spent: str # in hours
    recorded: bool = False
    row_index: int = -1  # 1-based row number in Excel

GREEN_HEX = "92D050" # Excel green used to mark "recorded"
GREEN_FILL = PatternFill(start_color=GREEN_HEX, end_color=GREEN_HEX, fill_type="solid")

def is_green(cell):
    """Returns True if the cell is filled with green (meaning: already recorded)."""
    try:
        fill = cell.fill
        fg = fill.fgColor
        color = str(fg.rgb).upper() if fg and hasattr(fg, "rgb") else ""
        return color.endswith(GREEN_HEX)
    except:
        return False

def DH_parse_timesheet(filepath: str) -> List[TimeEntry]:
    wb = load_workbook(filepath, data_only=False)
    ws = wb.active

    headers = [cell.value.strip().lower() if cell.value else "" for cell in ws[1]]
    col_index = {name: idx for idx, name in enumerate(headers)}

    required = ["date", "client", "matter", "description", "time (hours)"]
    for r in required:
        if r not in col_index:
            raise ValueError(f"Missing required column: '{r}'")

    entries = []
    current_date = None

    for i, row in enumerate(ws.iter_rows(min_row=4), start=4):  # start=4 to match Excel row numbers
        try:
            date_cell = row[col_index["date"]]
            date_val = date_cell.value

            if isinstance(date_val, datetime):
                current_date = date_val
            elif isinstance(date_val, str) and date_val.startswith("="):
                # formula cell, increment date
                if current_date is not None:
                    current_date = current_date + timedelta(days=1)
                else:
                    print(f"Skipping row {i} — no valid current_date to increment")
                    continue
            elif date_val is None:
                # blank cell, keep current_date as is
                if current_date is None:
                    print(f"Skipping row {i} — no current_date yet")
                    continue
            else:
                print(f"Skipping row {i} — unexpected date cell content: {date_val}")
                continue

            print(f"Row {i} — date resolved to: {current_date}")

            client_cell = row[col_index["client"]]
            matter_cell = row[col_index["matter"]]
            description_cell = row[col_index["description"]]
            time_cell = row[col_index["time (hours)"]]

            # --- Skip if marked recorded (green) ---
            if is_green(client_cell):
                continue

            # --- Extract rest of values ---
            client = str(client_cell.value).strip() if client_cell.value else ""
            if client.lower() == "administration":
                continue
            matter = str(matter_cell.value).strip() if matter_cell.value else ""
            description = str(description_cell.value).strip() if description_cell.value else ""
            time_spent = str(time_cell.value) if time_cell.value else ""

            if not client or not matter or not description or not time_spent:
                print(f"Skipping row {i} — missing required data.")
                continue

            entries.append(TimeEntry(
                date=current_date,
                client=client,
                matter=matter,
                description=description,
                time_spent=time_spent,
                row_index=i
            ))

        except Exception as e:
            print(f"Skipping row due to error: {e}")
            continue

    # Sort by date (just in case)
    entries.sort(key=lambda x: x.date)
    return entries

def DH_record_time_entry(entry: TimeEntry, filepath: str, confirm_before_saving=True):
    """Takes a TimeEntry object and initiates UI automation into PCLaw."""

    print(f"Recording time entry for {entry.matter}: {entry.client} on {entry.date} - {entry.time_spent}h")

    # Automation logic
    saved = DH_fill_time_entry(
        date=entry.date.strftime("%Y-%m-%d"),
        client=entry.client,
        matter=entry.matter,
        description=entry.description.replace("’", "'").replace("œ", "oe"),
        time_spent=entry.time_spent,
        confirm_before_saving=confirm_before_saving
    )

    if saved:
        entry.recorded = True
        DH_mark_entry_as_recorded(filepath, entry)

def DH_mark_entry_as_recorded(filepath: str, entry: TimeEntry):
    """Applies green fill to the 'client' cell of a recorded entry."""
    wb = load_workbook(filepath)
    ws = wb.active

    headers = [cell.value.strip().lower() if cell.value else "" for cell in ws[1]]
    col_index = {name: idx for idx, name in enumerate(headers)}
    client_col = col_index.get("client")

    if client_col is None:
        raise ValueError("Couldn't find 'client' column in Excel")

    cell = ws.cell(row=entry.row_index, column=client_col + 1)  # Excel is 1-based
    cell.fill = GREEN_FILL

    wb.save(filepath)
    alert_info(f"Marked entry for {entry.client} on {entry.date} as recorded.")
    print(f"Marked row {entry.row_index} as recorded.")
