from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

GREEN_HEX = "92D050"  # Excel green used to mark "recorded"

def is_green(cell):
    """Returns True if the cell has the 'recorded' green fill."""
    try:
        fill = cell.fill
        fg = fill.fgColor
        if not fg or not hasattr(fg, "rgb"):
            return False
        color = str(fg.rgb)  # force to string in case it's not
        return color.upper().endswith(GREEN_HEX)
    except:
        return False

def parse_DH_time_sheet(filepath):
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    # Get headers from row 1
    headers = [cell.value.strip().lower() if cell.value else "" for cell in ws[1]]
    col_index = {name: idx for idx, name in enumerate(headers)}

    # Ensure required columns are present
    required = ["date", "client", "description", "time (hours)"]
    for r in required:
        if r not in col_index:
            raise ValueError(f"Missing required column: '{r}'")

    entries = []
    current_date = None

    for row in ws.iter_rows(min_row=4):
        try:
            # Cells
            date_cell = row[col_index["date"]]
            client_cell = row[col_index["client"]]
            description_cell = row[col_index["description"]]
            time_cell = row[col_index["time (hours)"]]

            # Skip if client cell is green
            if is_green(client_cell):
                continue  # Skip recorded entry

            # Handle date
            date_raw = date_cell.value
            if isinstance(date_raw, datetime):
                current_date = date_raw.date()
            elif isinstance(date_raw, str):
                try:
                    current_date = datetime.strptime(date_raw.strip(), "%Y-%m-%d").date()
                except ValueError:
                    pass  # ignore bad string, don't override current_date

            if not current_date:
                continue  # Can't process row without any known date

            client = str(client_cell.value).strip() if client_cell.value else ""
            description = str(description_cell.value).strip() if description_cell.value else ""
            time_spent = float(time_cell.value) if time_cell.value else 0.0

            if not client or not description or not time_spent:
                continue

            entries.append({
                "date": current_date,
                "client": client,
                "description": description,
                "time_spent": time_spent,
                "recorded": False
            })

        except Exception as e:
            print(f"Skipping row due to error: {e}")
            continue

    return entries